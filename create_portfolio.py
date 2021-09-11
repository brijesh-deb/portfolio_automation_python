import openpyxl
import datetime
from openpyxl.styles import PatternFill, Font

# global variables
vr_mf_start_row=0
vr_mf_end_row=0
vr_curr_row=1

dilip_bk_ac = [{"Bank":"ICICI","A/C":"xxxxxxxxxxxxxxx","Amount":"0"},
               {"Bank":"SBI-Pension","A/C":"xxxxxxxxxxxxxxx","Amount":"0"},
               {"Bank":"SBI","A/C":"xxxxxxxxxxxxxxx","Amount":"0"},
               {"Bank":"Canara","A/C":"xxxxxxxxxxxxxxx","Amount":"0"}]

sobha_bk_ac = [{"Bank":"ICICI","A/C":"xxxxxxxxxxxxxxx","Amount":"0"},
               {"Bank":"SBI-Pension","A/C":"xxxxxxxxxxxxxxx","Amount":"0"}]

ratna_bk_ac = [{"Bank":"ICICI","A/C":"xxxxxxxxxxxxxxx","Amount":"0"},
               {"Bank":"HDFC","A/C":"xxxxxxxxxxxxxxx","Amount":"0"}]

brijesh_bk_ac = [{"Bank":"ICICI","A/C":"xxxxxxxxxxxxxxx","Amount":"0"},
               {"Bank":"ICICI","A/C":"xxxxxxxxxxxxxxx","Amount":"0"},
               {"Bank":"Canara","A/C":"xxxxxxxxxxxxxxx","Amount":"0"},
               {"Bank":"SBI","A/C":"xxxxxxxxxxxxxxx","Amount":"0"}]

prt_curr_row=1
prt_mf_tot_row = 0
prt_bank_tot_row= 0
prt_owner=""

file = openpyxl.load_workbook('my-investment-overviewxls-29-Aug-2021--2025.xltx')
ws = file.active
p_file = openpyxl.Workbook()
p_ws = p_file.active


# Main function which will orchestrate portfolio file creation
def create_portfolio():
    get_vr_cordinate()      # Get important row numbers from VR file
    mf_detail_lst= extract_mf_detail(vr_mf_start_row,vr_mf_end_row) # Extract MF details from VR file
    write_portfolio_header()        # Write Portfolio header
    write_mf_detail(mf_detail_lst)  # Write MF details in portfolio
    if prt_owner=="brijesh":
        write_stock_details()
    write_bank_detail()             # Write Bank details in Portfolio
    if prt_owner=="ratna" or prt_owner=="brijesh":
        write_retiral_detail()
    if prt_owner=="brijesh":
        write_realestate_details()
    write_networth()                # Write Networth in Portfolio
    write_networth_breakup()        # Write Networth breakup in Portfolio
    p_file.save("portfolio.xlsx")   # Save Portfolio file in local folder


def write_retiral_detail():
    pass

def write_stock_details():
    pass

def write_realestate_details():
    pass

def write_networth_breakup():
    add_blank_space()
    write_header(heading_lst=["Networth Breakup","Amount","VR%"])
    write_nw_brk_detail()


def add_blank_space():
    global prt_curr_row
    prt_curr_row +=1


def write_nw_brk_detail():
    global prt_curr_row, prt_bank_tot_row
    # Equity part of MF
    p_ws.cell(prt_curr_row,1).value="Equity - MF"
    p_ws.cell(prt_curr_row,2).value='=($F${0}*$C${1})/100'.format(prt_mf_tot_row,prt_curr_row)
    p_ws.cell(prt_curr_row,3).value=0
    mark_input_cell(prt_curr_row,3)
    prt_curr_row += 1
    # Debt part of MF
    p_ws.cell(prt_curr_row,1).value="Debt - MF"
    p_ws.cell(prt_curr_row,2).value='=($F${0}*$C${1})/100'.format(prt_mf_tot_row,prt_curr_row)
    p_ws.cell(prt_curr_row,3).value=0
    mark_input_cell(prt_curr_row,3)
    prt_curr_row += 1
    # Cash part of MF
    p_ws.cell(prt_curr_row,1).value="Cash - MF"
    p_ws.cell(prt_curr_row,2).value='=($F${0}*$C${1})/100'.format(prt_mf_tot_row,prt_curr_row)
    p_ws.cell(prt_curr_row,3).value=0
    mark_input_cell(prt_curr_row,3)
    prt_curr_row += 1
    # Other part of MF
    p_ws.cell(prt_curr_row,1).value="Other - MF"
    p_ws.cell(prt_curr_row,2).value='=($F${0}*$C${1})/100'.format(prt_mf_tot_row,prt_curr_row)
    p_ws.cell(prt_curr_row,3).value=0
    mark_input_cell(prt_curr_row,3)
    prt_curr_row += 1
    # Cash - Bank
    p_ws.cell(prt_curr_row,1).value="Cash - Bank"
    p_ws.cell(prt_curr_row,2).value='=$C{0}'.format(prt_bank_tot_row)
    p_ws.cell(prt_curr_row,3).value="NA"
    prt_curr_row += 1


def write_networth():
    global prt_curr_row
    yellowFill = PatternFill('solid', openpyxl.styles.colors.Color(rgb="FFFF66"))
    prt_curr_row += 1 # skip 1 row
    p_ws.cell(prt_curr_row,1).value="Networth"
    p_ws.cell(prt_curr_row,2).value= '=SUM($F${0},$C${1})'.format(prt_mf_tot_row,prt_bank_tot_row)   # Formula to calculate Networth
    for col in range(1,3):
        p_ws.cell(prt_curr_row,col).fill = yellowFill
        p_ws.cell(prt_curr_row,col).font = Font(bold=True)
    prt_curr_row += 1


def write_bank_detail():
    global prt_curr_row, prt_bank_tot_row
    add_blank_space()
    write_header(heading_lst=["Bank","A/C","Amt"])
    bank_detail_list=[]
    if prt_owner=="Dilip":
        bank_detail_list = dilip_bk_ac
    elif prt_owner=="Sobha":
        bank_detail_list = sobha_bk_ac
    elif prt_owner=="Ratna":
        bank_detail_list = ratna_bk_ac
    elif prt_owner=="Brijesh":
        bank_detail_list = brijesh_bk_ac
    bank_detail_start_row = prt_curr_row
    bank_detail_end_row = prt_curr_row + len(bank_detail_list)-1
    for item in bank_detail_list:
        p_ws.cell(prt_curr_row,1).value=item["Bank"]
        p_ws.cell(prt_curr_row,2).value=item["A/C"]
        p_ws.cell(prt_curr_row,3).value=item["Amount"]
        mark_input_cell(prt_curr_row,3)
        prt_curr_row += 1
    p_ws.cell(prt_curr_row,1).value="Bank Total"
    p_ws.cell(prt_curr_row,2).value=""
    p_ws.cell(prt_curr_row,3).value='=SUM($C${0}:$C${1})'.format(bank_detail_start_row,bank_detail_end_row)  # add bank total formula here

    make_row_bold(prt_curr_row)

    prt_bank_tot_row = prt_curr_row
    prt_curr_row += 1


def write_mf_detail(mf_detail_lst):
    global prt_curr_row, prt_mf_tot_row
    add_blank_space()
    write_header(heading_lst=["Fund Name","Folio","Units","Cost/Unit", "Tot Cost", "Curr Val","Tot Return", "NAV","CAGR","VR"])
    for mf_detail in mf_detail_lst:
        p_ws.cell(prt_curr_row, 1).value = mf_detail["Name"]
        p_ws.cell(prt_curr_row, 2).value = mf_detail["Folio"]
        p_ws.cell(prt_curr_row, 3).value = mf_detail["Units"]
        p_ws.cell(prt_curr_row, 4).value = mf_detail["Cost_Unit"]
        p_ws.cell(prt_curr_row, 5).value = mf_detail["Tot_Cost"]
        p_ws.cell(prt_curr_row, 6).value = mf_detail["Curr_Val"]
        p_ws.cell(prt_curr_row, 7).value = mf_detail["Tot_Return"]
        p_ws.cell(prt_curr_row, 8).value = mf_detail["NAV"]
        p_ws.cell(prt_curr_row, 9).value = mf_detail["CAGR"]
        p_ws.cell(prt_curr_row, 10).value = mf_detail["VR"]
        prt_curr_row +=1
    prt_mf_tot_row = prt_curr_row-1
    make_row_bold(prt_mf_tot_row)


# write portfolio header
def write_portfolio_header():
    global prt_curr_row
    p_ws.cell(prt_curr_row,1).value="Portfolio Owner"
    p_ws.cell(prt_curr_row,2).value=prt_owner
    p_ws.cell(prt_curr_row,3).value="Period"
    p_ws.cell(prt_curr_row,4).value=""
    p_ws.cell(prt_curr_row,5).value="Compiled On"
    p_ws.cell(prt_curr_row,6).value=str(datetime.date.today())
    mark_input_cell(prt_curr_row,4)
    make_row_bold(1)
    prt_curr_row += 1


def write_header(heading_lst):
    global prt_curr_row
    yellowFill = PatternFill('solid', openpyxl.styles.colors.Color(rgb="FFFF66"))
    heading_index=1
    for heading in heading_lst:
        p_ws.cell(prt_curr_row, heading_index).value = heading
        p_ws.cell(prt_curr_row,heading_index).fill = yellowFill
        p_ws.cell(prt_curr_row,heading_index).font = Font(bold=True)
        heading_index +=1
    prt_curr_row += 1

def mark_input_cell(row,col):
    blueFill = PatternFill('solid', openpyxl.styles.colors.Color(rgb="99FFFF"))
    p_ws.cell(row, col).fill = blueFill

def make_row_bold(row):
    for cell in p_ws['${0}:{1}'.format(row,row)]:
        cell.font = Font(bold=True)

def extract_mf_detail(vr_mf_start_row,vr_mf_end_row):
    mf_detail_lst=[]
    for row in range(vr_mf_start_row,vr_mf_end_row):
        mf_detail = {}
        if float(ws.cell(row,11).value) >1:
            mf_detail["Name"]=ws.cell(row,1).value
            mf_detail["Folio"]=ws.cell(row,2).value
            mf_detail["Units"]=ws.cell(row,11).value
            mf_detail["Cost_Unit"]=ws.cell(row,9).value
            mf_detail["Tot_Cost"]=ws.cell(row,8).value
            mf_detail["Curr_Val"]=ws.cell(row,10).value
            mf_detail["Tot_Return"]=ws.cell(row,13).value
            mf_detail["NAV"]=ws.cell(row,5).value
            mf_detail["CAGR"]=ws.cell(row,14).value
            mf_detail["VR"]=ws.cell(row,3).value
            mf_detail_lst.append(mf_detail)
    #add MF total row
    mf_detail = {}
    row = vr_mf_end_row+1
    mf_detail["Name"] = ws.cell(row, 1).value
    mf_detail["Folio"] = ws.cell(row, 2).value
    mf_detail["Units"] = ws.cell(row, 11).value
    mf_detail["Cost_Unit"] = ws.cell(row, 9).value
    mf_detail["Tot_Cost"] = ws.cell(row, 8).value
    mf_detail["Curr_Val"] = ws.cell(row, 10).value
    mf_detail["Tot_Return"] = ws.cell(row, 13).value
    mf_detail["NAV"] = ws.cell(row, 5).value
    mf_detail["CAGR"] = ws.cell(row, 14).value
    mf_detail["VR"] = ws.cell(row, 3).value
    mf_detail_lst.append(mf_detail)

    return mf_detail_lst


# Get co-ordinates of details in VR File
def get_vr_cordinate():
    global vr_curr_row, vr_mf_start_row, vr_mf_end_row,prt_owner
    for row in ws.rows:
        if str(row[0].value).strip() == "Fund Name":
            vr_mf_start_row=vr_curr_row + 1
        elif str(row[0].value).strip() == "Mutual Funds":
            vr_mf_end_row = vr_curr_row-1
        elif str(row[0].value).strip() == "Selected labels":
            prt_owner=row[2].value

        vr_curr_row += 1


create_portfolio()